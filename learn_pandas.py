import pandas as pd
from openpyxl import load_workbook

filePath ="C:/Users/shdbt/Desktop/DATAFILE/조방/"
fileName = 'policeStation.xlsx'
sheetName = 'DATA'

## paramter : 
# 경로 : 파일 경로
# engine : 일반적으로 excel을 불러오기 위해서는 openpyxl engine 사용
# sheet_name : 시트 이름
# header : 몇 번째 열에서 시작되는지
# index_col : 특정 열을 인덱스로 지정하고 싶을 때 ex) index_col = '도로명주소'
# usecols : 특정 열들만 골라서 dataframe 으로 만든다. ex) 'C,D,E'
df = pd.read_excel(filePath+fileName, engine='openpyxl', sheet_name=sheetName, usecols='C')
df2 = pd.read_excel(filePath+fileName, engine='openpyxl', sheet_name=sheetName, usecols='D')
address =[]

for i in df2.values:
    address.append(str(i))

for i in address :
    print(i)

# w=load_workbook(filePath + fileName,read_only=True)
# ws = w[sheetName]
# row =2
# column =3
# for i in range(2000):
#     print(ws.cell(row,column).value)
#     row += 1